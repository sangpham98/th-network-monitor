from datetime import UTC, datetime, timedelta
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app import auth
from app.config import settings
from app.database import Base, get_db
from app.main import app
from app.models import Incident, Store


def utc_now() -> datetime:
    return datetime(2026, 5, 5, 8, 30, 0, tzinfo=UTC).replace(tzinfo=None)


def configure_auth(monkeypatch):
    monkeypatch.setattr(settings, "auth_enabled", True)
    monkeypatch.setattr(settings, "admin_username", "admin")
    monkeypatch.setattr(settings, "admin_password", "secret")
    monkeypatch.setattr(settings, "session_secret", "test-secret")
    monkeypatch.setattr(settings, "session_cookie_name", "test_session")
    monkeypatch.setattr(settings, "session_max_age_seconds", 28800)


def make_db_override():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)
    db = SessionLocal()
    store_1 = Store(store_code="CH001", pc_name="PC001", region="HCM", area="Area A")
    store_2 = Store(store_code="CH002", pc_name="PC002", region="HN", area="Area B")
    db.add_all([store_1, store_2])
    db.commit()
    db.refresh(store_1)
    db.refresh(store_2)
    now = utc_now()
    db.add_all(
        [
            Incident(
                store_id=store_1.id,
                incident_type="DOWN",
                status="OPEN",
                started_at=now,
                alert_sent=True,
                detail="Store down",
            ),
            Incident(
                store_id=store_2.id,
                incident_type="WAN_DOWN",
                status="RESOLVED",
                started_at=now - timedelta(days=1),
                ended_at=now,
                duration_seconds=3600,
                recovery_sent=True,
                detail="WAN recovered",
            ),
        ]
    )
    db.commit()

    def override_get_db():
        try:
            yield db
        finally:
            pass

    app.dependency_overrides[get_db] = override_get_db
    return store_1, store_2


def clear_overrides():
    app.dependency_overrides.clear()


def authed_client():
    client = TestClient(app)
    client.cookies.set("test_session", auth.create_session_token("admin"))
    return client


def workbook_rows(content: bytes):
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Incidents"]
    return list(sheet.iter_rows(values_only=True))


def test_incident_export_redirects_to_login_when_unauthenticated(monkeypatch):
    configure_auth(monkeypatch)
    make_db_override()
    client = TestClient(app)

    response = client.get("/incidents/export", follow_redirects=False)

    clear_overrides()
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_incident_export_returns_xlsx(monkeypatch):
    configure_auth(monkeypatch)
    make_db_override()
    client = authed_client()

    response = client.get("/incidents/export")

    clear_overrides()
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "incident_report_" in response.headers["content-disposition"]


def test_incident_export_workbook_contains_headers_and_rows(monkeypatch):
    configure_auth(monkeypatch)
    make_db_override()
    client = authed_client()

    response = client.get("/incidents/export")
    rows = workbook_rows(response.content)

    clear_overrides()
    assert rows[0][:4] == ("Incident ID", "Store Code", "PC Name", "Region")
    assert any(row[1] == "CH001" and row[5] == "DOWN" for row in rows[1:])
    assert any(row[1] == "CH002" and row[6] == "RESOLVED" for row in rows[1:])


def test_incident_export_status_filter(monkeypatch):
    configure_auth(monkeypatch)
    make_db_override()
    client = authed_client()

    response = client.get("/incidents/export?status=OPEN")
    rows = workbook_rows(response.content)

    clear_overrides()
    assert len(rows) == 2
    assert rows[1][6] == "OPEN"
    assert rows[1][1] == "CH001"


def test_incident_export_store_code_filter(monkeypatch):
    configure_auth(monkeypatch)
    make_db_override()
    client = authed_client()

    response = client.get("/incidents/export?store_code=CH002")
    rows = workbook_rows(response.content)

    clear_overrides()
    assert len(rows) == 2
    assert rows[1][1] == "CH002"


def test_incidents_page_has_export_and_store_links(monkeypatch):
    configure_auth(monkeypatch)
    store_1, _store_2 = make_db_override()
    client = authed_client()

    response = client.get("/incidents?status=OPEN&store_code=CH001")

    clear_overrides()
    assert response.status_code == 200
    assert "/incidents/export?status=OPEN&amp;store_code=CH001" in response.text
    assert f'href="/stores/{store_1.id}"' in response.text
    assert "2026-05-05 15:30:00" in response.text
    assert "CH002" not in response.text
