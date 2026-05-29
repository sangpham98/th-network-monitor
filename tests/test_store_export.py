from datetime import UTC, datetime, timedelta
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app import auth
from app.config import settings
from app.database import Base, get_db
from app.main import app, build_store_timelines
from app.models import Incident, Store, StoreStatus


def utc_now() -> datetime:
    return datetime(2026, 5, 5, 8, 30, 0, tzinfo=UTC).replace(tzinfo=None)


def configure_auth(monkeypatch):
    monkeypatch.setattr(settings, "auth_enabled", True)
    monkeypatch.setattr(settings, "admin_username", "admin")
    monkeypatch.setattr(settings, "admin_password", "secret")
    monkeypatch.setattr(settings, "session_secret", "test-secret")
    monkeypatch.setattr(settings, "session_cookie_name", "test_session")
    monkeypatch.setattr(settings, "session_max_age_seconds", 28800)


def make_db_override():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)
    db = SessionLocal()
    store_1 = Store(
        store_code="70000123",
        pc_name="PC001",
        region="HCM",
        area="Area A",
        address="123 Test Street",
        wan_dns="wan.example",
        ip_tunnel="10.0.0.1",
        ip_local="192.168.1.10",
        enabled=True,
    )
    store_2 = Store(
        store_code="70000124",
        pc_name="PC002",
        region="HN",
        area="Area B",
        address="456 Test Street",
        wan_dns="wan2.example",
        ip_tunnel="10.0.0.2",
        ip_local="192.168.1.11",
        enabled=False,
    )
    db.add_all([store_1, store_2])
    db.commit()
    db.refresh(store_1)
    db.refresh(store_2)
    now = utc_now()
    db.add_all(
        [
            StoreStatus(
                store_id=store_1.id,
                wan_status="UP",
                tunnel_status="DOWN",
                overall_status="TUNNEL_DOWN",
                last_check_at=now,
            ),
            StoreStatus(
                store_id=store_2.id,
                wan_status="UP",
                tunnel_status="UP",
                overall_status="UP",
                last_check_at=now,
            ),
        ]
    )
    db.commit()

    def override_get_db():
        try:
            yield db
        finally:
            pass

    app.dependency_overrides[get_db] = override_get_db
    return store_1, store_2


def clear_overrides():
    app.dependency_overrides.clear()


def authed_client():
    client = TestClient(app)
    client.cookies.set("test_session", auth.create_session_token("admin"))
    return client


def workbook_rows(content: bytes):
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Stores"]
    return list(sheet.iter_rows(values_only=True))


def test_store_export_redirects_to_login_when_unauthenticated(monkeypatch):
    configure_auth(monkeypatch)
    make_db_override()
    client = TestClient(app)

    response = client.get("/stores/export", follow_redirects=False)

    clear_overrides()
    assert response.status_code == 303
    assert response.headers["location"] == "/login"


def test_store_export_returns_xlsx(monkeypatch):
    configure_auth(monkeypatch)
    make_db_override()
    client = authed_client()

    response = client.get("/stores/export")

    clear_overrides()
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "store_report_" in response.headers["content-disposition"]


def test_store_export_workbook_contains_headers_and_rows(monkeypatch):
    configure_auth(monkeypatch)
    make_db_override()
    client = authed_client()

    response = client.get("/stores/export")
    rows = workbook_rows(response.content)

    clear_overrides()
    assert rows[0] == ("Mã CH", "PC Name", "IP Local", "IP Tunnel", "WAN DNS", "Miền", "Khu vực", "Địa chỉ")
    assert any(row[0] == "70000123" and row[2] == "192.168.1.10" and row[4] == "wan.example" for row in rows[1:])
    assert any(row[0] == "70000124" and row[5] == "HN" and row[7] == "456 Test Street" for row in rows[1:])


def test_store_export_q_filter(monkeypatch):
    configure_auth(monkeypatch)
    make_db_override()
    client = authed_client()

    response = client.get("/stores/export?q=PC002")
    rows = workbook_rows(response.content)

    clear_overrides()
    assert len(rows) == 2
    assert rows[1][0] == "70000124"


def test_store_export_status_filter(monkeypatch):
    configure_auth(monkeypatch)
    make_db_override()
    client = authed_client()

    response = client.get("/stores/export?status=TUNNEL_DOWN")
    rows = workbook_rows(response.content)

    clear_overrides()
    assert len(rows) == 2
    assert rows[1][0] == "70000123"
    assert rows[1][3] == "10.0.0.1"


def test_stores_page_has_export_link_and_no_check_now(monkeypatch):
    configure_auth(monkeypatch)
    make_db_override()
    client = authed_client()

    response = client.get("/stores?q=PC001&status=TUNNEL_DOWN")

    clear_overrides()
    assert response.status_code == 200
    assert "/stores/export?q=PC001&amp;status=TUNNEL_DOWN" in response.text
    assert "Export Excel" in response.text
    assert "Stability History" in response.text
    assert "00:00 → 24:00" in response.text
    assert "Check now" not in response.text
    assert 'action="/monitor/run-once"' not in response.text


def test_build_store_timelines_clips_incidents_to_current_day(monkeypatch):
    monkeypatch.setattr(settings, "timezone", "Asia/Ho_Chi_Minh")
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)
    db = SessionLocal()
    store = Store(store_code="70000123", pc_name="PC001")
    db.add(store)
    db.commit()
    db.refresh(store)
    now = datetime(2026, 5, 5, 12, 0, 0, tzinfo=UTC)
    db.add(
        Incident(
            store_id=store.id,
            incident_type="DOWN",
            status="OPEN",
            started_at=(now - timedelta(hours=2)).replace(tzinfo=None),
        )
    )
    db.commit()

    timelines = build_store_timelines(db, [store], now)

    clear_overrides()
    assert len(timelines[store.id]) == 1
    segment = timelines[store.id][0]
    assert segment["status"] == "DOWN"
    assert 70.8 <= segment["left"] <= 70.9
    assert 8.3 <= segment["width"] <= 8.4
    assert "DOWN 17:00→19:00" == segment["label"]
